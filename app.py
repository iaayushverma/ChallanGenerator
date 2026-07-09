"""
Automated Challan Generator - Backend API
Client: Timeless Impressions Printers
"""

import os
import sys
import sqlite3
import datetime
import shutil
import webview
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from xml.sax.saxutils import escape

def get_base_path():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def fmt_num(val):
    if val is None or val == "":
        return ""
    try:
        f = float(val)
        if f.is_integer():
            return str(int(f))
        return f"{f:.2f}"
    except (ValueError, TypeError):
        return str(val).strip()

class ChalaanAPI:
    def __init__(self):
        self.db_path = os.path.join(os.path.expanduser("~"), "ChalaanApp_Data.db")
        self.init_db()
        self.current_estimate_data = {}
        self.current_store_master = {}
        self.estimate_store_names = {}

    def init_db(self):
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            c.execute('''CREATE TABLE IF NOT EXISTS chalaan_seq (id INTEGER PRIMARY KEY, seq INTEGER)''')
            c.execute('''CREATE TABLE IF NOT EXISTS employees (id INTEGER PRIMARY KEY, name TEXT UNIQUE)''')
            c.execute('''CREATE TABLE IF NOT EXISTS chalaans (
                            chalaan_no INTEGER PRIMARY KEY,
                            store_code TEXT,
                            store_name TEXT,
                            campaign_name TEXT,
                            state TEXT,
                            date TEXT,
                            employee_name TEXT,
                            status TEXT
                         )''')
            
            c.execute("SELECT COUNT(*) FROM chalaan_seq")
            if c.fetchone()[0] == 0:
                c.execute("INSERT INTO chalaan_seq (seq) VALUES (100)")
            conn.commit()


    def get_next_chalaan_no(self):
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            c.execute("SELECT seq FROM chalaan_seq WHERE id = 1")
            seq = c.fetchone()[0]
            c.execute("UPDATE chalaan_seq SET seq = seq + 1 WHERE id = 1")
            conn.commit()
            return seq

    def select_file(self):
        window = webview.windows[0]
        result = window.create_file_dialog(webview.FileDialog.OPEN, allow_multiple=False, file_types=('Excel Files (*.xlsx)',))
        return result[0] if result else None

    def select_folder(self):
        window = webview.windows[0]
        result = window.create_file_dialog(webview.FileDialog.FOLDER)
        return result[0] if result else None

    def find_col_index(self, headers, possible_names):
        for i, header in enumerate(headers):
            if not header: continue
            clean_header = " ".join(str(header).split()).lower().replace('"', '').replace('*', '')
            for name in possible_names:
                if name in clean_header:
                    return i
        return -1

    def load_master_file(self, filepath):
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            idx_code = self.find_col_index(headers, ['store code'])
            idx_name = self.find_col_index(headers, ['store name'])
            
            if idx_code == -1 or idx_name == -1:
                return {"error": "Could not find 'Store Code' or 'Store Name' in Master file headers."}
            
            self.current_store_master = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[idx_code]
                if code is not None:
                    clean_code = str(code).strip().upper()
                    name_val = row[idx_name]
                    clean_name = str(name_val).strip() if name_val is not None else ""
                    
                    if clean_name != "":
                        self.current_store_master[clean_code] = clean_name
                    elif clean_code not in self.current_store_master:
                        self.current_store_master[clean_code] = ""
                        
            return {"success": True, "count": len(self.current_store_master)}
        except Exception as e:
            return {"error": str(e)}

    def load_estimate_file(self, filepath):
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            mapping = {
                'store_code': self.find_col_index(headers, ['store code', 'store']),
                'est_store_name': self.find_col_index(headers, ['store name']), 
                'campaign': self.find_col_index(headers, ['campaign name']),
                'particular': self.find_col_index(headers, ['article descript']),
                'w': self.find_col_index(headers, ['"w"', ' w ', '^w$']),
                'h': self.find_col_index(headers, ['"h"', ' h ', '^h$']),
                'sft': self.find_col_index(headers, ['sft', 'sq ft']),
                'qty': self.find_col_index(headers, ['qty', 'quantity'])
            }
            
            if mapping['w'] == -1: mapping['w'] = self.find_col_index(headers, ['w'])
            if mapping['h'] == -1: mapping['h'] = self.find_col_index(headers, ['h'])

            missing = [k for k, v in mapping.items() if v == -1 and k != 'est_store_name']
            if missing: return {"error": f"Missing required columns in Estimate: {', '.join(missing)}"}

            self.current_estimate_data = {}
            self.estimate_store_names = {}
            unmatched = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_code = row[mapping['store_code']]
                if not raw_code: continue
                
                s_code = str(raw_code).strip().upper()

                if mapping['est_store_name'] != -1 and row[mapping['est_store_name']]:
                    self.estimate_store_names[s_code] = str(row[mapping['est_store_name']]).strip()

                item = {
                    'campaign': row[mapping['campaign']],
                    'particular': row[mapping['particular']],
                    'w': row[mapping['w']],
                    'h': row[mapping['h']],
                    'sft': row[mapping['sft']],
                    'qty': row[mapping['qty']]
                }
                
                if s_code not in self.current_store_master:
                    unmatched.add(s_code)
                
                if s_code not in self.current_estimate_data:
                    self.current_estimate_data[s_code] = []
                self.current_estimate_data[s_code].append(item)

            matched_keys = list(set(self.current_estimate_data.keys()) - unmatched)
            return {
                "success": True, 
                "matched_count": len(matched_keys),
                "unmatched": list(unmatched),
                "stores": matched_keys
            }
        except Exception as e:
            return {"error": str(e)}

    def generate_pdfs(self, stores_to_process, state, output_dir, client_name="Reliance"):
        if not output_dir or not os.path.exists(output_dir):
            return {"error": "Invalid or no save location selected."}

        generated_count = 0
        today_date = datetime.datetime.now().strftime("%d.%m.%Y")
        
        logo_path = os.path.join(get_base_path(), "assets", "logo.png")
        sign_path = os.path.join(get_base_path(), "assets", "sign.png")
        
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=6, leading=8)

        # NEW: Determine if we are in Bulk Mode
        is_bulk = len(stores_to_process) > 1
        elements = []
        
        # NEW: Setup a single massive Document for Bulk Mode
        if is_bulk:
            timestamp = datetime.datetime.now().strftime("%H%M%S")
            safe_state = state.replace(" ", "_") if state else "UnknownState"
            bulk_filename = os.path.join(output_dir, f"Bulk_Challans_{safe_state}_{today_date.replace('.','-')}_{timestamp}.pdf")
            doc = SimpleDocTemplate(bulk_filename, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

        for store_code in stores_to_process:
            items = self.current_estimate_data.get(store_code, [])
            if not items: continue

            # NEW: Insert a page break if we are in bulk mode and it's not the first store
            if is_bulk and generated_count > 0:
                elements.append(PageBreak())

            master_name = self.current_store_master.get(store_code, "")
            backup_name = self.estimate_store_names.get(store_code, "Unknown Store")
            final_store_name = master_name if master_name else backup_name

            chalaan_no = self.get_next_chalaan_no()
            
            if os.path.exists(logo_path):
                elements.append(Image(logo_path, width=530, height=80, kind='proportional', hAlign='CENTER'))
                elements.append(Spacer(1, 30))
            
            elements.append(Paragraph("<b>DELIVERY CHALLAN</b>", ParagraphStyle(name='Title', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=18, alignment=1, spaceAfter=28)))
            
            safe_store_name = escape(str(final_store_name))
            safe_client_name = escape(str(client_name))
            
            header_data = [
                ["Client:", Paragraph(safe_client_name, styles['Normal']), "Dated:", today_date],
                ["Store Address:", Paragraph(safe_store_name, styles['Normal']), "Challan no.:", str(chalaan_no)],
                ["Store Code:", store_code, "State:", state]
            ]
            
            htable = Table(header_data, colWidths=[80, 200, 80, 100])
            htable.setStyle(TableStyle([
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
                ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ]))
            elements.append(htable)
            elements.append(Spacer(1, 15))
            
            table_data = [["Sr. No.", "Campaign Name", "Particular", "Size-W", "Size-H", "Sq ft", "Qty"]]
            total_qty = 0
            
            for idx, item in enumerate(items, 1):
                try:
                    q_val = float(item['qty']) if item['qty'] else 0
                    total_qty += int(q_val) if q_val.is_integer() else q_val
                except ValueError:
                    pass
                
                safe_camp = escape(str(item['campaign'] or ''))
                safe_part = escape(str(item['particular'] or ''))
                
                camp_para = Paragraph(safe_camp, cell_style)
                part_para = Paragraph(safe_part, cell_style)
                
                table_data.append([
                    str(idx),
                    camp_para,
                    part_para,
                    fmt_num(item['w']),
                    fmt_num(item['h']),
                    fmt_num(item['sft']),
                    fmt_num(item['qty'])
                ])
            
            table_data.append(["", "", "", "", "", "Total:", fmt_num(total_qty)])
                
            item_table = Table(table_data, colWidths=[35, 120, 180, 50, 50, 50, 50])
            t_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('INNERGRID', (0, 0), (-1, -2), 0.25, colors.black), 
                ('BOX', (0, 0), (-1, -2), 0.25, colors.black),       
                ('FONTNAME', (5, -1), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (5, -1), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, -1), (-1, -1), 10),
            ])
            item_table.setStyle(t_style)
            elements.append(item_table)
            elements.append(Spacer(1, 40))
            
            if os.path.exists(sign_path):
                elements.append(Image(sign_path, width=100, height=40, kind='proportional', hAlign='RIGHT'))
            elements.append(Paragraph("<b>Authorized Signatory</b>", ParagraphStyle(name='Sign', parent=styles['Normal'], alignment=2)))
            
            # DB Insert            
            # Compress all campaigns into a single string for tracking
            campaigns = list(set([str(item['campaign']).strip() for item in items if item['campaign']]))
            campaign_str = ", ".join(campaigns)[:100] 

            with sqlite3.connect(self.db_path, timeout=10) as conn:
                c = conn.cursor()
                c.execute("INSERT INTO chalaans (chalaan_no, store_code, store_name, state, date, employee_name, status, campaign_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                          (chalaan_no, store_code, final_store_name, state, today_date, "Unassigned", "Pending", campaign_str))
                conn.commit()
                
            generated_count += 1
            
            # NEW: If we are generating just a single store, build the single PDF here and clear out the elements
            if not is_bulk:
                pdf_filename = os.path.join(output_dir, f"Challan_{chalaan_no}_{store_code}.pdf")
                single_doc = SimpleDocTemplate(pdf_filename, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                single_doc.build(elements)
                elements = []

        # NEW: If we are in Bulk Mode, compile all pages into the single massive document
        if is_bulk and generated_count > 0:
            doc.build(elements)

        return {"success": True, "count": generated_count, "path": output_dir}

    def get_employees(self):
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            c.execute("SELECT name FROM employees")
            emps = [row[0] for row in c.fetchall()]
            return emps

    def add_employee(self, name):
        try:
            with sqlite3.connect(self.db_path, timeout=10) as conn:
                c = conn.cursor()
                c.execute("INSERT INTO employees (name) VALUES (?)", (name,))
                conn.commit()
                return {"success": True}
        except sqlite3.IntegrityError:
            return {"error": "Employee already exists"}

    def get_tracking_data(self):
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            c.execute("SELECT chalaan_no, store_code, store_name, campaign_name, date, employee_name, status FROM chalaans ORDER BY chalaan_no DESC")
            data = [{"chalaan_no": r[0], "store_code": r[1], "store_name": r[2], "campaign_name": r[3] or "", "date": r[4], "employee": r[5], "status": r[6]} for r in c.fetchall()]
            return data

    def update_tracking(self, chalaan_no, employee, status):
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            c.execute("UPDATE chalaans SET employee_name = ?, status = ? WHERE chalaan_no = ?", (employee, status, chalaan_no))
            conn.commit()
            return {"success": True}
    
    def delete_tracking(self, chalaan_no):
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            c.execute("DELETE FROM chalaans WHERE chalaan_no = ?", (chalaan_no,))
            conn.commit()
            return {"success": True}

        # ================= BACKUP SYSTEM =================

    def export_backup(self):
        window = webview.windows[0]
        # Open a Save dialog asking where they want to store the backup
        result = window.create_file_dialog(webview.FileDialog.SAVE, save_filename='ChallanApp_Data_Backup.db')
        if result:
            try:
                # Copy the live database file to their chosen location
                shutil.copy2(self.db_path, result[0])
                return {"success": True, "path": result[0]}
            except Exception as e:
                return {"error": str(e)}
        return {"error": "Export cancelled."}

    def import_backup(self):
        window = webview.windows[0]
        # Open an Open dialog to select an old backup file
        result = window.create_file_dialog(webview.FileDialog.OPEN, file_types=('Database Files (*.db)',))
        if result:
            try:
                # Security check: Make sure it's actually our app's database before overwriting
                test_conn = sqlite3.connect(result[0])
                c = test_conn.cursor()
                c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='chalaans'")
                if not c.fetchone():
                    return {"error": "Invalid file. This does not appear to be a valid Challan backup database."}
                test_conn.close()
                
                # Overwrite the live database with the backup
                shutil.copy2(result[0], self.db_path)
                return {"success": True}
            except Exception as e:
                return {"error": f"Failed to restore backup: {str(e)}"}
        return {"error": "Import cancelled."}
    
    def bulk_update_tracking(self, chalaan_nos, employee, status):
        if not chalaan_nos: 
            return {"success": False, "error": "No chalaans selected."}
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            placeholders = ','.join(['?'] * len(chalaan_nos))
            
            updates = []
            params = []
            
            # Only update the fields the user actually selected
            if employee:
                updates.append("employee_name = ?")
                params.append(employee)
            if status:
                updates.append("status = ?")
                params.append(status)
            
            if not updates:
                return {"success": True}
            
            query = f"UPDATE chalaans SET {', '.join(updates)} WHERE chalaan_no IN ({placeholders})"
            params.extend(chalaan_nos)
            
            c.execute(query, params)
            conn.commit()
            return {"success": True}

    def bulk_delete_tracking(self, chalaan_nos):
        if not chalaan_nos: 
            return {"success": False, "error": "No chalaans selected."}
        with sqlite3.connect(self.db_path, timeout=10) as conn:
            c = conn.cursor()
            placeholders = ','.join(['?'] * len(chalaan_nos))
            c.execute(f"DELETE FROM chalaans WHERE chalaan_no IN ({placeholders})", chalaan_nos)
            conn.commit()
            return {"success": True}


if __name__ == '__main__':
    api = ChalaanAPI()
    template_path = os.path.join(get_base_path(), 'gui', 'index.html')
    webview.create_window('Timeless Impressions - Challan Generator', template_path, js_api=api, width=1050, height=800)
    webview.start()